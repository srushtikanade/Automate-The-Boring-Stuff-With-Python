import openpyxl

myfile='example.xlsx'
wb =openpyxl.Workbook(myfile)
sheet=wb.active
row_count = sheet.max_row
column_count = sheet.max_column
content=[]
for row in range(len(row_count+1)):
    for cell in range(len(column_count+1)):
        celldata=sheet.cell(row=row,column=cell).value
        content.append(celldata)

wbnew =openpyxl.Workbook()
sheetnew=wbnew.active
print('Inverting spreadsheet...')

for row in range(1,row_count+1):
    for cell in range(1,column_count+1):
        sheet.cell(row=cell,column=row).value=content[row-1][cell-1]
wbnew.save('Inverted-'+myfile)
print('Inverted your spreadsheet and saved with an inverted prefix!')

