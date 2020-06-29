import openpyxl,sys

#def blankRowInserter(start,nBlanks,)
if len(sys.argv)>3:
    startCell=sys.argv[1]
    nblanks=sys.argv[2]
    myfile=sys.argv[3]


wb =openpyxl.Workbook(myfile)
sheet=wb.active
row_count = sheet.max_row
column_count = sheet.max_column
content=[]
for row in range(len(row_count+1)):
    for cell in range(len(column_count)):
        celldata=sheet.cell(row=row,column=cell).value
        content.append(celldata)

wbnew =openpyxl.Workbook()
sheetnew=wbnew.active
print('Inserting blanks...')

for row in range(1,start):
    for cell in range(1,column_count+1):
        sheet.cell(row=row,column=cell).value=content[row-1][cell-1]

for row in range(start+nblanks,nblanks+row_count+1):
    for cell in range(1,column_count+1):
        sheet.cell(row=row,column=cell).value=content[row-nblanks-1][cell-1]
wbnew.save('BlankRowInserted-'+myfile)
        

