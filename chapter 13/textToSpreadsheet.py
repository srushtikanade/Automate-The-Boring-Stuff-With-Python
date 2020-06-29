import openpyxl
from openpyxl.styles import Font
from openpyxl.utils import get_column_letter


filepath=input('Enter the absolute path to the folder where the files are saved:')

files=input('Enter the file names to insert into spreadsheet :') #seperated by space
filesList = files.split()


wb = openpyxl.Workbook()
sheet = wb.active

columnNum = 1
for file in filesList:
    
    lines = open(filepath + '\\' + file).readlines()

    toBold = Font(bold=True)
    sheet.cell(row=1, column=columnNum).value = file
    sheet.cell(row=1, column=columnNum).font = toBold

    longest = 0
    row_num = 2
    for line in lines:
        line = line.strip()

        
        if len(line) > longest:
            longest = len(line)

        
        sheet.cell(row=row_num, column=columnNum).value = line
        row_num += 1

    column_letter = get_column_letter(columnNum)
    sheet.column_dimensions[column_letter].width = longest
    column_num += 1

wb.save(filepath + '\\text2sheet'+file)
print('Spreadsheet saved with text2sheet prefix in the same directory')

