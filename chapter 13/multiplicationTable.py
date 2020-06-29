import openpyxl
from openpyxl.styles import Font

def multiplicationTable(number):
    wb =openpyxl.Workbook()
    sheet=wb.active
    sheet.freeze_panes = 'A2'
    sheet.freeze_panes = 'B1'
    fontObj1 = Font(name='Times New Roman', bold=True)
    for i in range(1,number+1):
        sheet.cell(row=i+1,column=1).value=i
        sheet.cell(row=i+1,column=1).font=fontObj1
    for i in range(1,number+1):
        sheet.cell(row=1,column=i+1).value=i
        sheet.cell(row=1,column=i+1).font=fontObj1
    for row in range(1,number+1):
        for cell in range(1,number+1):
            sheet.cell(row=row+1,column=cell+1).value=row*cell
    wb.save('multiplicationTable'+str(number)+'.xlsx')
        
myNumber=int(input('Enter the number you want to create a multiplication table for :'))     
multiplicationTable(myNumber)

