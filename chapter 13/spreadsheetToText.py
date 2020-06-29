import openpyxl,os
#os.chdir(r'C:\Users\yourName\OneDrive\Documents')
# add path if you arent in that directory currently or chdir

print('Enter the file names to insert into spreadsheet :') #seperated by space
      
myExcelFile = input()

wb=openpyxl.load_workbook(myExcelFile)
sheet=wb.active
maxRow=sheet.max_row
maxCol=sheet.max_column

content=[]
columnNum=1
filename=sheet.cell(row=1,column=columnNum).value
print('Reading column'+str(columnNum)+' of excel file.')
for row in range(2,maxRow+1):
    if sheet.cell(row=row,column=columnNum).value!=None: # this is an important condition or else you will end up copying blanks cells to you text file
        cellData=sheet.cell(row=row,column=columnNum).value
        content.append(cellData)
print('Reading that column'+str(columnNum)+' data to text file.')
myfile=open(str(columnNum)+str(filename)+'sheet2text.txt','w')
for line in content:
    myfile.write(line+'\n')
myfile.close()
print('Saved that column to text file with columnNum and Text2sheet as prefix')

columnNum+=1
wb.close()
    
