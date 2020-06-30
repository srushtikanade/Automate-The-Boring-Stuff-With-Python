import openpyxl,os,csv


for file in os.listdir('.'):
    
    if file.endswith('.xlsx'):
        wb=openpyxl.load_workbook(file)
        for sheetName in wb.get_sheet_names():
                
            sheet=wb.get_sheet_by_name(sheetName)
            FileName=file[:-5] # .xlsx is 5 char extension, prior to it is the name of file

            csvFile=open(FileName+'-'+sheetName+'.csv','w',newline='')
            csvWriter=csv.writer(csvFile)

            for rowNum in range(1,sheet.max_row+1):
                rowData=[]
                for columnNum in range(1,sheet.max_column+1):
                    rowData.append(sheet.cell(row=rowNum,column=columnNum).value)
                for row in rowData:
                    csvWriter.writerow(row)
            print('Converted excel file'+FileName +'to csv')
            csvFile.close()

